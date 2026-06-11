#!/usr/bin/env python3
"""
Convertitore Excel -> JSON per il sito statico Totomondiale.

Uso rapido:
  python convert_excel_to_json.py
  python convert_excel_to_json.py "../TotoMondiale Risposte.xlsx"

Lo script riconosce la struttura del Google Form gia compilato e genera:
  data/partecipanti.json
  data/partite.json
  data/classifica.json
  data/regolamento.json
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Modulo mancante: openpyxl. Installa con: python -m pip install openpyxl"
    ) from exc


ARROW = "\u2192"

TEAM_DISPLAY_NAMES = {
    "ARABIA SAUDITA": "Arabia Saudita",
    "ARGENTINA": "Argentina",
    "ALGERIA": "Algeria",
    "AUSTRALIA": "Australia",
    "AUSTRIA": "Austria",
    "BELGIO": "Belgio",
    "BOSNIA": "Bosnia",
    "BRASILE": "Brasile",
    "CANADA": "Canada",
    "CAPO VERDE": "Capo Verde",
    "COLOMBIA": "Colombia",
    "COREA DEL SUD": "Corea del Sud",
    "COSTA D'AVORIO": "Costa d'Avorio",
    "CROAZIA": "Croazia",
    "CURAZAO": "Curazao",
    "ECUADOR": "Ecuador",
    "EGITTO": "Egitto",
    "FRANCIA": "Francia",
    "GERMANIA": "Germania",
    "GHANA": "Ghana",
    "GIAPPONE": "Giappone",
    "GIORDANIA": "Giordania",
    "HAITI": "Haiti",
    "INGHILTERRA": "Inghilterra",
    "IRAN": "Iran",
    "IRAQ": "Iraq",
    "MAROCCO": "Marocco",
    "MESSICO": "Messico",
    "NORVEGIA": "Norvegia",
    "NUOVA ZELANDA": "Nuova Zelanda",
    "PAESI BASSI": "Paesi Bassi",
    "PANAMA": "Panama",
    "PARAGUAY": "Paraguay",
    "PORTOGALLO": "Portogallo",
    "QATAR": "Qatar",
    "RD CONGO": "RD Congo",
    "REP. CECA": "Rep. Ceca",
    "SCOZIA": "Scozia",
    "SENEGAL": "Senegal",
    "SPAGNA": "Spagna",
    "SUDAFRICA": "Sudafrica",
    "SVEZIA": "Svezia",
    "SVIZZERA": "Svizzera",
    "TUNISIA": "Tunisia",
    "TURCHIA": "Turchia",
    "URUGUAY": "Uruguay",
    "USA": "USA",
    "UZBEKISTAN": "Uzbekistan",
}

BONUS_FINALI = {
    "bonus_vincitore": ("vincitore", "Vincitore mondiale"),
    "bonus_capocannoniere": ("capocannoniere", "Capocannoniere"),
    "bonus_assistman": ("assistman", "Miglior assistman"),
    "bonus_portiere": ("portiere", "Miglior portiere"),
    "bonus_giocatore": ("giocatore", "Miglior giocatore"),
    "bonus_difesa": ("difesa", "Miglior difesa"),
    "bonus_attacco": ("attacco", "Miglior attacco"),
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def strip_symbols(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in value if unicodedata.category(ch)[0] not in {"S", "C"})


def slugify(value: str, fallback: str = "item") -> str:
    raw = unicodedata.normalize("NFKD", value)
    raw = raw.encode("ascii", "ignore").decode("ascii")
    raw = re.sub(r"[^a-zA-Z0-9]+", "-", raw.lower()).strip("-")
    return raw or fallback


def normalize_lookup(value: Any) -> str:
    text = clean_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def display_team(value: str) -> str:
    cleaned = clean_text(value).upper()
    if cleaned in TEAM_DISPLAY_NAMES:
        return TEAM_DISPLAY_NAMES[cleaned]
    words = []
    for word in cleaned.split():
        if word in {"USA", "RD"}:
            words.append(word)
        elif word in {"DEL", "DELLA", "D", "DI"}:
            words.append(word.lower())
        else:
            words.append(word.capitalize())
    return " ".join(words)


def normalize_outcome(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if int(value) == value and int(value) in {1, 2}:
            return str(int(value))
    text = clean_text(value).upper().replace(".0", "")
    if text in {"1", "X", "2"}:
        return text
    return text or None


def parse_score(value: Any) -> tuple[int, int] | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return None
    text = clean_text(value)
    match = re.search(r"(\d+)\D+(\d+)", text)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def normalize_score(value: Any) -> str | None:
    score = parse_score(value)
    if score:
        return f"{score[0]}-{score[1]}"
    text = clean_text(value)
    return text or None


def outcome_from_score(score: tuple[int, int] | None) -> str | None:
    if not score:
        return None
    if score[0] > score[1]:
        return "1"
    if score[0] < score[1]:
        return "2"
    return "X"


def cell_value(cell: Any) -> Any:
    value = cell.value
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def classify_header(header: Any) -> str:
    text = clean_text(header)
    upper = text.upper()
    if "INFORMAZIONI CRONOLOGICHE" in upper:
        return "timestamp"
    if "NICKNAME" in upper:
        return "partecipante"
    if "PRONOSTICO" in upper and (ARROW in text or "->" in text):
        return "pronostico_1x2"
    if "RISULTATO ESATTO" in upper and (ARROW in text or "->" in text):
        return "pronostico_risultato_esatto"
    if "PRIMA QUALIFICATA" in upper:
        return "bonus_prima_qualificata_gruppo"
    if "SECONDA QUALIFICATA" in upper:
        return "bonus_seconda_qualificata_gruppo"
    if "VINCITORE" in upper or "CAMPIONE" in upper:
        return "bonus_vincitore"
    if "FINALISTA" in upper:
        return "bonus_finalista"
    if "CAPOCANNONIERE" in upper:
        return "bonus_capocannoniere"
    if "ASSISTMAN" in upper:
        return "bonus_assistman"
    if "PORTIERE" in upper:
        return "bonus_portiere"
    if "GIOCATORE" in upper:
        return "bonus_giocatore"
    if "DIFESA" in upper:
        return "bonus_difesa"
    if "ATTACCO" in upper:
        return "bonus_attacco"
    return "altro"


def subject_from_header(header: Any) -> str:
    text = clean_text(header)
    if ARROW in text:
        return text.split(ARROW, 1)[1].strip()
    if "->" in text:
        return text.split("->", 1)[1].strip()
    group = re.search(r"GRUPPO\s+([A-Z])", text.upper())
    if group:
        return group.group(1)
    label = strip_symbols(text)
    return clean_text(label)


def split_match(subject: str) -> tuple[str, str]:
    parts = re.split(r"\s+VS\s+", subject, flags=re.IGNORECASE)
    if len(parts) != 2:
        return display_team(subject), ""
    return display_team(parts[0]), display_team(parts[1])


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return deepcopy(default)
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return deepcopy(default)


def default_regolamento(generated_at: str, latest_excel_update: str | None) -> dict[str, Any]:
    return {
        "generatedAt": generated_at,
        "ultimoAggiornamento": latest_excel_update or generated_at,
        "fonte": "Regolamento modificabile: i punteggi non erano presenti nel file Excel.",
        "titolo": "Regolamento Totomondiale",
        "descrizione": (
            "Modifica questo file per cambiare testi e punteggi. "
            "La classifica viene calcolata dal sito usando questi valori."
        ),
        "punteggi": {
            "risultatoEsatto": 3,
            "segnoCorretto": 1,
            "primaQualificata": 2,
            "secondaQualificata": 2,
            "vincitore": 10,
            "finalista": 5,
            "capocannoniere": 5,
            "assistman": 4,
            "portiere": 4,
            "giocatore": 5,
            "difesa": 3,
            "attacco": 3,
        },
        "regole": [
            "Risultato esatto: 3 punti.",
            "Segno 1/X/2 corretto senza risultato esatto: 1 punto.",
            "Partite senza risultato reale: 0 punti e stato da giocare.",
            "Bonus gruppi e bonus finali vengono assegnati quando inserisci i valori reali in data/partite.json.",
        ],
    }


def preserve_match_fields(match: dict[str, Any], existing: dict[str, Any] | None) -> dict[str, Any]:
    preserved = deepcopy(match)
    if not existing:
        return preserved
    for key in [
        "data",
        "ora",
        "stadio",
        "risultatoReale",
        "golCasa",
        "golTrasferta",
        "stato",
        "note",
    ]:
        if key in existing:
            preserved[key] = existing[key]
    return preserved


def match_actual_score(match: dict[str, Any]) -> tuple[int, int] | None:
    home_goals = match.get("golCasa")
    away_goals = match.get("golTrasferta")
    if isinstance(home_goals, int) and isinstance(away_goals, int):
        return home_goals, away_goals
    parsed = parse_score(match.get("risultatoReale"))
    if parsed:
        return parsed
    return None


def score_match(
    prediction: dict[str, Any], match: dict[str, Any], points: dict[str, int]
) -> dict[str, Any]:
    actual_score = match_actual_score(match)
    predicted_score = parse_score(prediction.get("risultatoEsatto"))
    predicted_outcome = prediction.get("pronostico") or outcome_from_score(predicted_score)

    base = {
        "partitaId": prediction["partitaId"],
        "gruppo": prediction.get("gruppo"),
        "partita": prediction.get("partita"),
        "pronostico": predicted_outcome,
        "risultatoEsatto": prediction.get("risultatoEsatto"),
        "risultatoReale": match.get("risultatoReale"),
        "stato": "da_giocare",
        "punti": 0,
    }
    if not actual_score:
        return base

    actual_outcome = outcome_from_score(actual_score)
    base["risultatoReale"] = f"{actual_score[0]}-{actual_score[1]}"
    if predicted_score == actual_score:
        base["stato"] = "corretto"
        base["punti"] = int(points.get("risultatoEsatto", 3))
    elif predicted_outcome and predicted_outcome == actual_outcome:
        base["stato"] = "parziale"
        base["punti"] = int(points.get("segnoCorretto", 1))
    else:
        base["stato"] = "sbagliato"
    return base


def score_bonus(
    participant: dict[str, Any],
    partite_data: dict[str, Any],
    points: dict[str, int],
) -> list[dict[str, Any]]:
    details: list[dict[str, Any]] = []
    real_groups = partite_data.get("qualificateReali", {})
    real_final = partite_data.get("bonusReali", {})
    predicted_groups = participant.get("bonus", {}).get("gruppi", {})
    predicted_final = participant.get("bonus", {}).get("finali", {})

    for group, values in predicted_groups.items():
        real_values = real_groups.get(group, {})
        for position, label, point_key in [
            ("prima", "Prima qualificata", "primaQualificata"),
            ("seconda", "Seconda qualificata", "secondaQualificata"),
        ]:
            predicted_value = values.get(position)
            real_value = real_values.get(position)
            status = "da_giocare"
            awarded = 0
            if real_value:
                if normalize_lookup(predicted_value) == normalize_lookup(real_value):
                    status = "corretto"
                    awarded = int(points.get(point_key, 0))
                else:
                    status = "sbagliato"
            details.append(
                {
                    "id": f"gruppo-{group.lower()}-{position}",
                    "tipo": point_key,
                    "label": f"{label} gruppo {group}",
                    "pronostico": predicted_value,
                    "valoreReale": real_value,
                    "stato": status,
                    "punti": awarded,
                }
            )

    for key, label in [
        ("vincitore", "Vincitore mondiale"),
        ("capocannoniere", "Capocannoniere"),
        ("assistman", "Miglior assistman"),
        ("portiere", "Miglior portiere"),
        ("giocatore", "Miglior giocatore"),
        ("difesa", "Miglior difesa"),
        ("attacco", "Miglior attacco"),
    ]:
        predicted_value = predicted_final.get(key)
        real_value = real_final.get(key)
        status = "da_giocare"
        awarded = 0
        if real_value:
            if normalize_lookup(predicted_value) == normalize_lookup(real_value):
                status = "corretto"
                awarded = int(points.get(key, 0))
            else:
                status = "sbagliato"
        details.append(
            {
                "id": key,
                "tipo": key,
                "label": label,
                "pronostico": predicted_value,
                "valoreReale": real_value,
                "stato": status,
                "punti": awarded,
            }
        )

    real_finalists = real_final.get("finalisti") or []
    if isinstance(real_finalists, str):
        real_finalists = [real_finalists]
    normalized_finalists = {normalize_lookup(item) for item in real_finalists}
    for key, label in [("finalista_1", "Finalista 1"), ("finalista_2", "Finalista 2")]:
        predicted_value = predicted_final.get(key)
        status = "da_giocare"
        awarded = 0
        if normalized_finalists:
            if normalize_lookup(predicted_value) in normalized_finalists:
                status = "corretto"
                awarded = int(points.get("finalista", 0))
            else:
                status = "sbagliato"
        details.append(
            {
                "id": key,
                "tipo": "finalista",
                "label": label,
                "pronostico": predicted_value,
                "valoreReale": real_finalists,
                "stato": status,
                "punti": awarded,
            }
        )
    return details


def build_classifica(
    partecipanti_data: dict[str, Any],
    partite_data: dict[str, Any],
    regolamento: dict[str, Any],
    generated_at: str,
) -> dict[str, Any]:
    matches_by_id = {item["id"]: item for item in partite_data.get("partite", [])}
    points = regolamento.get("punteggi", {})
    rows = []

    for participant in partecipanti_data.get("partecipanti", []):
        match_details = [
            score_match(prediction, matches_by_id.get(prediction["partitaId"], {}), points)
            for prediction in participant.get("pronostici", [])
        ]
        bonus_details = score_bonus(participant, partite_data, points)
        match_points = sum(item["punti"] for item in match_details)
        bonus_points = sum(item["punti"] for item in bonus_details)
        counts = Counter(item["stato"] for item in match_details)
        rows.append(
            {
                "partecipanteId": participant["id"],
                "nome": participant["nome"],
                "puntiTotali": match_points + bonus_points,
                "puntiPartite": match_points,
                "puntiBonus": bonus_points,
                "risultatiEsatti": counts.get("corretto", 0),
                "risultatiParziali": counts.get("parziale", 0),
                "risultatiSbagliati": counts.get("sbagliato", 0),
                "partiteDaGiocare": counts.get("da_giocare", 0),
                "dettaglio": {
                    "partite": match_details,
                    "bonus": bonus_details,
                },
            }
        )

    rows.sort(key=lambda item: (-item["puntiTotali"], item["nome"].lower()))
    for index, row in enumerate(rows, start=1):
        row["posizione"] = index

    return {
        "generatedAt": generated_at,
        "calcolataDa": "convert_excel_to_json.py",
        "classifica": rows,
    }


def build_data(excel_path: Path, data_dir: Path, sheet_name: str | None) -> dict[str, Any]:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    column_info = []
    for index, header in enumerate(headers, start=1):
        column_info.append(
            {
                "index": index,
                "type": classify_header(header),
                "subject": subject_from_header(header),
                "header": clean_text(header),
            }
        )

    existing_partite = load_json(data_dir / "partite.json", {})
    existing_by_id = {item.get("id"): item for item in existing_partite.get("partite", [])}
    existing_by_teams = {
        f"{normalize_lookup(item.get('casa'))}|{normalize_lookup(item.get('trasferta'))}": item
        for item in existing_partite.get("partite", [])
    }

    matches = []
    match_columns = []
    for column in column_info:
        if column["type"] != "pronostico_1x2":
            continue
        exact = next(
            (
                candidate
                for candidate in column_info
                if candidate["type"] == "pronostico_risultato_esatto"
                and candidate["subject"] == column["subject"]
            ),
            None,
        )
        order = len(matches) + 1
        group_index = (order - 1) // 6
        group = chr(ord("A") + group_index)
        home, away = split_match(column["subject"])
        match_id = f"gruppo-{group.lower()}-{((order - 1) % 6) + 1:02d}"
        base_match = {
            "id": match_id,
            "ordine": order,
            "fase": "Gironi",
            "gruppo": group,
            "casa": home,
            "trasferta": away,
            "partita": f"{home} vs {away}",
            "data": None,
            "ora": None,
            "stadio": None,
            "risultatoReale": None,
            "golCasa": None,
            "golTrasferta": None,
            "stato": "da_giocare",
            "note": "",
        }
        key = f"{normalize_lookup(home)}|{normalize_lookup(away)}"
        existing = existing_by_id.get(match_id) or existing_by_teams.get(key)
        match = preserve_match_fields(base_match, existing)
        if match_actual_score(match):
            match["stato"] = "giocata"
            match["risultatoReale"] = match["risultatoReale"] or (
                f"{match['golCasa']}-{match['golTrasferta']}"
                if isinstance(match.get("golCasa"), int)
                and isinstance(match.get("golTrasferta"), int)
                else None
            )
        matches.append(match)
        match_columns.append(
            {
                "id": match_id,
                "group": group,
                "subject": column["subject"],
                "predictionColumn": column["index"],
                "scoreColumn": exact["index"] if exact else None,
                "home": home,
                "away": away,
            }
        )

    group_bonus_columns: dict[str, dict[str, int]] = {}
    final_bonus_columns: dict[str, int] = {}
    finalista_counter = 0
    for column in column_info:
        if column["type"] == "bonus_prima_qualificata_gruppo":
            group_bonus_columns.setdefault(column["subject"], {})["prima"] = column["index"]
        elif column["type"] == "bonus_seconda_qualificata_gruppo":
            group_bonus_columns.setdefault(column["subject"], {})["seconda"] = column["index"]
        elif column["type"] == "bonus_finalista":
            finalista_counter += 1
            final_bonus_columns[f"finalista_{finalista_counter}"] = column["index"]
        elif column["type"] in BONUS_FINALI:
            final_bonus_columns[BONUS_FINALI[column["type"]][0]] = column["index"]

    latest_timestamp = None
    participants = []
    used_ids = Counter()
    for row in range(2, ws.max_row + 1):
        name = clean_text(cell_value(ws.cell(row, 2)))
        if not name:
            continue
        timestamp = cell_value(ws.cell(row, 1))
        latest_timestamp = max(
            [value for value in [latest_timestamp, timestamp] if value],
            default=timestamp,
        )
        base_id = slugify(name, fallback=f"partecipante-{row}")
        used_ids[base_id] += 1
        participant_id = base_id if used_ids[base_id] == 1 else f"{base_id}-{row}"

        predictions = []
        for match_column in match_columns:
            prediction_value = normalize_outcome(
                cell_value(ws.cell(row, match_column["predictionColumn"]))
            )
            score_value = normalize_score(
                cell_value(ws.cell(row, match_column["scoreColumn"]))
                if match_column["scoreColumn"]
                else None
            )
            predictions.append(
                {
                    "partitaId": match_column["id"],
                    "gruppo": match_column["group"],
                    "partita": f"{match_column['home']} vs {match_column['away']}",
                    "casa": match_column["home"],
                    "trasferta": match_column["away"],
                    "pronostico": prediction_value,
                    "risultatoEsatto": score_value,
                }
            )

        group_bonus = {}
        for group, positions in group_bonus_columns.items():
            group_bonus[group] = {
                "prima": clean_text(cell_value(ws.cell(row, positions.get("prima", 0))))
                if positions.get("prima")
                else None,
                "seconda": clean_text(cell_value(ws.cell(row, positions.get("seconda", 0))))
                if positions.get("seconda")
                else None,
            }

        final_bonus = {}
        for key, column_index in final_bonus_columns.items():
            final_bonus[key] = clean_text(cell_value(ws.cell(row, column_index))) or None

        participants.append(
            {
                "id": participant_id,
                "nome": name,
                "timestamp": timestamp,
                "sourceRow": row,
                "pronostici": predictions,
                "bonus": {
                    "gruppi": group_bonus,
                    "finali": final_bonus,
                },
            }
        )

    qualificate_reali = existing_partite.get("qualificateReali") or {
        chr(ord("A") + index): {"prima": None, "seconda": None} for index in range(12)
    }
    bonus_reali = existing_partite.get("bonusReali") or {
        "vincitore": None,
        "finalisti": [],
        "capocannoniere": None,
        "assistman": None,
        "portiere": None,
        "giocatore": None,
        "difesa": None,
        "attacco": None,
    }

    metadata = {
        "generatedAt": generated_at,
        "ultimoAggiornamento": latest_timestamp,
        "sourceFile": excel_path.name,
        "sheet": ws.title,
        "formulaCount": sum(
            1
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ),
    }

    partecipanti_data = {
        **metadata,
        "total": len(participants),
        "partecipanti": participants,
    }
    partite_data = {
        **metadata,
        "total": len(matches),
        "partite": matches,
        "qualificateReali": qualificate_reali,
        "bonusReali": bonus_reali,
    }

    regolamento = load_json(
        data_dir / "regolamento.json",
        default_regolamento(generated_at, latest_timestamp),
    )
    regolamento["generatedAt"] = generated_at
    regolamento.setdefault("ultimoAggiornamento", latest_timestamp or generated_at)

    classifica_data = build_classifica(
        partecipanti_data, partite_data, regolamento, generated_at
    )
    return {
        "partecipanti": partecipanti_data,
        "partite": partite_data,
        "regolamento": regolamento,
        "classifica": classifica_data,
        "summary": {
            "sheet": ws.title,
            "participants": len(participants),
            "matches": len(matches),
            "formulas": metadata["formulaCount"],
        },
    }


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    project_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Converte Excel Totomondiale in JSON.")
    parser.add_argument(
        "excel",
        nargs="?",
        default=str(project_dir.parent / "TotoMondiale Risposte.xlsx"),
        help="Percorso del file Excel sorgente.",
    )
    parser.add_argument(
        "--data-dir",
        default=str(project_dir / "data"),
        help="Cartella di output dei JSON.",
    )
    parser.add_argument("--sheet", default=None, help="Nome del foglio da leggere.")
    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    data_dir = Path(args.data_dir).resolve()
    if not excel_path.exists():
        raise SystemExit(f"File Excel non trovato: {excel_path}")
    data_dir.mkdir(parents=True, exist_ok=True)

    output = build_data(excel_path, data_dir, args.sheet)
    write_json(data_dir / "partecipanti.json", output["partecipanti"])
    write_json(data_dir / "partite.json", output["partite"])
    write_json(data_dir / "regolamento.json", output["regolamento"])
    write_json(data_dir / "classifica.json", output["classifica"])

    print(
        "Conversione completata: "
        f"{output['summary']['participants']} partecipanti, "
        f"{output['summary']['matches']} partite, "
        f"{output['summary']['formulas']} formule."
    )
    print(f"JSON salvati in: {data_dir}")


if __name__ == "__main__":
    main()
